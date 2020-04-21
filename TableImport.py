import selenium,time
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

company_name = "PSX"

print("Company name is : ", company_name)

chromedriver = "D:\Sublime\Projects/chromedriver"
browser = webdriver.Chrome(chromedriver)
#browser.maximize_window()
browser.get("https://dps.psx.com.pk/historical")

symbol = browser.find_element_by_id('historicalSymbolSearch')

symbol.clear()
symbol.send_keys(company_name)

search_sign = browser.find_element_by_id('historicalSymbolSearchAutocompleteList')
search_sign.click()
time.sleep(1)


#months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
#years = ['2016', '2017', '2018', '2019']
years = [2020]
months = ['March']

i = 0
z = 3
filename = "StockMarket.xlsx"
wb = load_workbook(filename)
wb.create_sheet(company_name)
#ws = wb.worksheets[0]        
sh = wb[company_name]

c = sh.cell(1,1)# = row_value
c.value = company_name

c1_1 = sh.cell(2,1)
c1_1.value = "Time"

c1_2 = sh.cell(2,2)
c1_2.value = "Open"

c1_3 = sh.cell(2,3)
c1_3.value = "High"

c1_4 = sh.cell(2,4)
c1_4.value = "Low"

c1_5 = sh.cell(2,5)
c1_5.value = "Close"

c1_6 = sh.cell(2,6)
c1_6.value = "Volume"

wb.save(filename)

for num1 in years:
    y = str(years[i])
    j = 0
    for num2 in months:
        m = str(months[j])
        link = "//select[@name='sector']/option[text()='"+m+"']"
        browser.find_element_by_xpath(link).click()

        link1 = "//select[@name='sector']/option[text()='"+y+"']"
        browser.find_element_by_xpath(link1).click()

        search_button = browser.find_element_by_id('historicalSymbolBtn')
        search_button.click()
        time.sleep(3)

        wb = load_workbook(filename)
        #ws = wb.worksheets[0]
        sh = wb[company_name]
        entries = browser.find_element_by_xpath(link)
        found = False

        while found != True:
            try: 
                entries = browser.find_element_by_xpath("//*[@id='historicalTable_info']")
                #found == True
                if (entries == browser.find_element_by_xpath("//*[@id='historicalTable_info']")):
                    found == True
            except Exception:
                if (entries == browser.find_element_by_xpath("//*[@id='historicalTable_info']")):
                    found == True

        entries_int = int(entries.text[13]+entries.text[14])
        coloumn_count = 7
        
        for a in range(1,entries_int+1):
            
            for b in range(1,coloumn_count):
             
                #print(type(a))
                d = int(b)
                row_value = browser.find_element_by_xpath("//*[@id='historicalTable']/tbody/tr["+str(a)+"]/td["+str(b)+"]").text
                                    
                c1 = sh.cell(int(z),int(b))# = row_value
                c1.value = row_value

            z+=1
                
        wb.save(filename)
        
        j += 1
    i+=1

browser.close()